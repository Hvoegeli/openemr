"""Clinical workbook (.xlsx) parser — deterministic, no LLM call.

Workbooks ship as a four-sheet dashboard:

- **Patient** — key/value demographics + allergies + As_Of_Date
- **Medications** — one row per active rx, columns:
  Brand | Generic | Strength | Route | Sig | Indication |
  Start_Date | Last_Filled | Refills_Remaining | Prescriber
- **Labs_Trend** — one row per test, fixed columns
  (Test | LOINC | Units | Reference_Range) followed by N date columns,
  each holding the value drawn on that day.
- **Care_Gaps** — one row per HEDIS/USPSTF measure, columns:
  Measure | HEDIS_or_USPSTF_ref | Status | Last_Done | Due_Date | Notes

The parser is hand-rolled on top of openpyxl. It tolerates header-name
variants (case + underscore variations) so a workbook generated from a
slightly different template still maps correctly. Any sheet missing
from the workbook is silently skipped — the resulting `Workbook` just
has empty list fields for the missing sheets.

Citations are sheet/row indexed. `bbox=None` everywhere — xlsx has no
spatial layout to overlay against.
"""

from __future__ import annotations

import io
import logging
from datetime import date, datetime

import openpyxl

from app.extraction.schemas import (
    Citation,
    Workbook,
    WorkbookCareGap,
    WorkbookCareGapStatus,
    WorkbookLabTrend,
    WorkbookLabValue,
    WorkbookMedication,
)

log = logging.getLogger("agent.extraction.workbook")


class WorkbookParseError(ValueError):
    """Raised when a workbook can't be loaded as a valid .xlsx file or
    the Patient sheet's structure is unrecognized. Subclass of ValueError
    so FastAPI maps it to a 400 the same way other render-side errors
    are handled."""


# ──────────────────────────────────────────────────────────────────────────
# Cell coercion helpers — openpyxl returns Python types when cells are
# typed; we normalize to the small set the schema expects.
# ──────────────────────────────────────────────────────────────────────────

def _str_or_none(value: object) -> str | None:
    """Return a stripped non-empty string or None. Numbers stringify."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _date_or_none(value: object) -> date | None:
    """Coerce a cell value to a date, or None if the cell isn't datelike.

    openpyxl returns `datetime.datetime` for date-typed cells when the
    cell stores a real serial date, and a string for cells where the
    user typed text like '2024-10-18'. We accept both."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # Try ISO first (workbooks generated programmatically use it),
        # then a couple of US-style variants. Anything else returns None.
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _int_or_none(value: object) -> int | None:
    """Coerce to int, returning None for blanks or unparsable strings."""
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        try:
            return int(value)
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            return None
    return None


def _slug(s: str) -> str:
    """Lowercase + collapse non-alphanumerics to a single underscore.
    Used to build snake_case `field_or_chunk_id` slugs from headers."""
    out: list[str] = []
    last_under = False
    for ch in s.lower():
        if ch.isalnum():
            out.append(ch)
            last_under = False
        elif not last_under:
            out.append("_")
            last_under = True
    return "".join(out).strip("_") or "row"


# ──────────────────────────────────────────────────────────────────────────
# Citation builder
# ──────────────────────────────────────────────────────────────────────────

def _wb_citation(
    *,
    source_document_id: str,
    page_or_section: str,
    field_or_chunk_id: str,
    quote_or_value: str,
) -> Citation:
    """Build a Citation rooted at a workbook sheet/row. `bbox=None`."""
    return Citation(
        source_type="workbook",
        source_id=source_document_id,
        page_or_section=page_or_section,
        field_or_chunk_id=field_or_chunk_id,
        quote_or_value=quote_or_value,
        bbox=None,
    )


# ──────────────────────────────────────────────────────────────────────────
# Per-sheet parsers
# ──────────────────────────────────────────────────────────────────────────

# Patient-sheet field-name aliases. Workbook templates vary in how they
# spell the labels (`DOB` vs `Date of Birth`, `MRN` vs `Medical Record
# Number`). We accept any of these for each canonical bucket.
_PATIENT_FIELD_ALIASES: dict[str, str] = {
    # canonical: matched aliases (lowercased, space-stripped)
    "name":          "patient_name",
    "patientname":   "patient_name",
    "fullname":      "patient_name",
    "dob":           "patient_dob",
    "dateofbirth":   "patient_dob",
    "birthdate":     "patient_dob",
    "mrn":           "patient_mrn",
    "medicalrecord": "patient_mrn",
    "pcp":           "pcp_name",
    "pcpname":       "pcp_name",
    "primarycare":   "pcp_name",
    "insurance":     "insurance",
    "payer":         "insurance",
    "allergies":     "allergies_text",
    "knownallergies": "allergies_text",
    "asofdate":      "as_of_date",
    "asof":          "as_of_date",
    "snapshotdate":  "as_of_date",
}


def _parse_patient_sheet(ws, *, source_document_id: str) -> dict:
    """Walk the Patient sheet's key/value rows and return a partial
    Workbook field dict. Unknown labels are silently ignored — the
    schema's optional fields default to None.
    """
    out: dict[str, object] = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 2 or row[0] is None:
            continue
        label = str(row[0]).strip().lower().replace(" ", "").replace("_", "")
        if label in ("field", "label"):
            continue  # header row
        canonical = _PATIENT_FIELD_ALIASES.get(label)
        if canonical is None:
            continue
        raw_value = row[1]
        if canonical in ("patient_dob", "as_of_date"):
            out[canonical] = _date_or_none(raw_value)
        else:
            out[canonical] = _str_or_none(raw_value)
    # _date_or_none / _str_or_none return None for empty cells; only
    # surface the keys that actually had a value so the schema sees
    # `None` for cells the workbook omitted entirely vs. cells where
    # the value happened to be empty.
    return {k: v for k, v in out.items() if v is not None}


def _row_dict(headers: list[str], row: tuple) -> dict:
    """Map a tuple row onto its header-keyed dict. Header names are
    lowercased + underscore-normalized for forgiving lookup."""
    out: dict[str, object] = {}
    for h, v in zip(headers, row, strict=False):
        if h is None:
            continue
        key = str(h).strip().lower().replace(" ", "_")
        out[key] = v
    return out


def _parse_medications_sheet(ws, *, source_document_id: str) -> list[WorkbookMedication]:
    """Walk the Medications sheet. Row 1 is headers; subsequent rows
    are one medication each. Empty rows (every cell None) are skipped."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[WorkbookMedication] = []
    for row_idx, row_tuple in enumerate(rows[1:], start=2):
        if all(c is None for c in row_tuple):
            continue
        cells = _row_dict(headers, row_tuple)
        brand = _str_or_none(cells.get("brand"))
        generic = _str_or_none(cells.get("generic"))
        if not brand and not generic:
            log.info("Medications row %d: skipping (no brand/generic name)", row_idx)
            continue
        label = generic or brand or f"row_{row_idx}"
        med = WorkbookMedication(
            brand=brand,
            generic=generic,
            strength=_str_or_none(cells.get("strength")),
            route=_str_or_none(cells.get("route")),
            sig=_str_or_none(cells.get("sig")),
            indication=_str_or_none(cells.get("indication")),
            start_date=_date_or_none(cells.get("start_date")),
            last_filled=_date_or_none(cells.get("last_filled")),
            refills_remaining=_int_or_none(cells.get("refills_remaining")),
            prescriber=_str_or_none(cells.get("prescriber")),
            source_citation=_wb_citation(
                source_document_id=source_document_id,
                page_or_section=f"Medications row {row_idx}",
                field_or_chunk_id=f"medications.row_{row_idx}.{_slug(label)}",
                quote_or_value=label,
            ),
        )
        out.append(med)
    return out


def _parse_labs_trend_sheet(ws, *, source_document_id: str) -> list[WorkbookLabTrend]:
    """Walk the Labs_Trend sheet.

    Row 1 mixes fixed columns (Test | LOINC | Units | Reference_Range)
    with N dynamic date columns. We detect which column indices are
    date-headers by attempting a date-coerce on each header cell;
    columns that aren't datelike fall into the "fixed" set.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
    header_row = rows[0]
    fixed_idx: dict[str, int] = {}
    date_columns: list[tuple[int, date]] = []
    for col_idx, raw in enumerate(header_row):
        as_date = _date_or_none(raw)
        if as_date is not None:
            date_columns.append((col_idx, as_date))
        elif raw is not None:
            key = str(raw).strip().lower().replace(" ", "_")
            fixed_idx[key] = col_idx

    if "test" not in fixed_idx:
        log.info("Labs_Trend: no 'Test' header found — skipping sheet")
        return []

    out: list[WorkbookLabTrend] = []
    for row_idx, row_tuple in enumerate(rows[1:], start=2):
        if all(c is None for c in row_tuple):
            continue
        test_name = _str_or_none(row_tuple[fixed_idx["test"]] if fixed_idx["test"] < len(row_tuple) else None)
        if not test_name:
            continue
        loinc = _str_or_none(row_tuple[fixed_idx["loinc"]]) if "loinc" in fixed_idx and fixed_idx["loinc"] < len(row_tuple) else None
        units = _str_or_none(row_tuple[fixed_idx["units"]]) if "units" in fixed_idx and fixed_idx["units"] < len(row_tuple) else None
        ref = _str_or_none(row_tuple[fixed_idx["reference_range"]]) if "reference_range" in fixed_idx and fixed_idx["reference_range"] < len(row_tuple) else None

        values: list[WorkbookLabValue] = []
        for col_idx, dcol in date_columns:
            cell = row_tuple[col_idx] if col_idx < len(row_tuple) else None
            cell_str = _str_or_none(cell)
            if cell_str is None:
                continue
            values.append(WorkbookLabValue(collection_date=dcol, value=cell_str))

        out.append(WorkbookLabTrend(
            test_name=test_name,
            loinc=loinc,
            units=units,
            reference_range=ref,
            values=values,
            source_citation=_wb_citation(
                source_document_id=source_document_id,
                page_or_section=f"Labs_Trend row {row_idx} ({test_name})",
                field_or_chunk_id=f"labs_trend.row_{row_idx}.{_slug(test_name)}",
                quote_or_value=test_name,
            ),
        ))
    return out


def _parse_care_gaps_sheet(ws, *, source_document_id: str) -> list[WorkbookCareGap]:
    """Walk the Care_Gaps sheet. Status values are normalized to the
    schema's accepted vocabulary; unknown statuses become None."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[WorkbookCareGap] = []
    accepted_statuses: set[str] = set(WorkbookCareGapStatus.__args__)  # type: ignore[attr-defined]
    for row_idx, row_tuple in enumerate(rows[1:], start=2):
        if all(c is None for c in row_tuple):
            continue
        cells = _row_dict(headers, row_tuple)
        measure = _str_or_none(cells.get("measure"))
        if not measure:
            continue
        raw_status = _str_or_none(cells.get("status")) or ""
        status: WorkbookCareGapStatus | None = None
        if raw_status.upper() in accepted_statuses:
            status = raw_status.upper()  # type: ignore[assignment]
        elif raw_status:
            log.info(
                "Care_Gaps row %d: unknown status %r (storing as None)",
                row_idx, raw_status,
            )
        out.append(WorkbookCareGap(
            measure=measure,
            reference=_str_or_none(cells.get("hedis_or_uspstf_ref")) or _str_or_none(cells.get("reference")),
            status=status,
            last_done=_date_or_none(cells.get("last_done")),
            due_date=_date_or_none(cells.get("due_date")),
            notes=_str_or_none(cells.get("notes")),
            source_citation=_wb_citation(
                source_document_id=source_document_id,
                page_or_section=f"Care_Gaps row {row_idx} ({measure})",
                field_or_chunk_id=f"care_gaps.row_{row_idx}.{_slug(measure)}",
                quote_or_value=measure,
            ),
        ))
    return out


# ──────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────

# Sheet-name aliases: the canonical names are Patient / Medications /
# Labs_Trend / Care_Gaps, but workbooks generated from older templates
# sometimes use spaces or singular forms. Lookup is case-insensitive and
# strips trailing/leading whitespace.
_SHEET_ALIASES: dict[str, str] = {
    "patient":     "patient",
    "demographics": "patient",
    "medications": "medications",
    "medication":  "medications",
    "meds":        "medications",
    "labs_trend":  "labs_trend",
    "labstrend":   "labs_trend",
    "labs":        "labs_trend",
    "care_gaps":   "care_gaps",
    "caregaps":    "care_gaps",
    "care gaps":   "care_gaps",
    "gaps":        "care_gaps",
}


def parse_workbook(
    file_bytes: bytes,
    *,
    source_document_id: str,
) -> Workbook:
    """Parse a .xlsx workbook into a typed `Workbook`.

    Args:
        file_bytes: Raw .xlsx bytes (as uploaded).
        source_document_id: The DocumentReference/{uuid} the source has
            been persisted under. Baked into every Citation.

    Returns:
        A fully-validated Workbook. Raises WorkbookParseError if the
        bytes don't load as a valid .xlsx, or if the result fails the
        Workbook schema check.
    """
    if not file_bytes:
        raise WorkbookParseError("workbook is empty")
    try:
        # `data_only=True` returns computed cell values rather than
        # formula strings. Workbooks generated from spreadsheets that
        # haven't been opened in Excel may have None for formula cells —
        # acceptable for our demo data which uses literal values.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise WorkbookParseError(f"failed to open workbook: {exc}") from exc

    sheets_by_canonical: dict[str, object] = {}
    for sheet_name in wb.sheetnames:
        canonical = _SHEET_ALIASES.get(sheet_name.strip().lower())
        if canonical is not None and canonical not in sheets_by_canonical:
            sheets_by_canonical[canonical] = wb[sheet_name]

    patient_fields: dict[str, object] = {}
    if "patient" in sheets_by_canonical:
        patient_fields = _parse_patient_sheet(
            sheets_by_canonical["patient"],
            source_document_id=source_document_id,
        )

    medications: list[WorkbookMedication] = []
    if "medications" in sheets_by_canonical:
        medications = _parse_medications_sheet(
            sheets_by_canonical["medications"],
            source_document_id=source_document_id,
        )

    lab_trends: list[WorkbookLabTrend] = []
    if "labs_trend" in sheets_by_canonical:
        lab_trends = _parse_labs_trend_sheet(
            sheets_by_canonical["labs_trend"],
            source_document_id=source_document_id,
        )

    care_gaps: list[WorkbookCareGap] = []
    if "care_gaps" in sheets_by_canonical:
        care_gaps = _parse_care_gaps_sheet(
            sheets_by_canonical["care_gaps"],
            source_document_id=source_document_id,
        )

    log.info(
        "parsed workbook: meds=%d lab_trends=%d care_gaps=%d source=%s",
        len(medications), len(lab_trends), len(care_gaps), source_document_id,
    )

    return Workbook(
        patient_name=patient_fields.get("patient_name"),  # type: ignore[arg-type]
        patient_dob=patient_fields.get("patient_dob"),  # type: ignore[arg-type]
        patient_mrn=patient_fields.get("patient_mrn"),  # type: ignore[arg-type]
        pcp_name=patient_fields.get("pcp_name"),  # type: ignore[arg-type]
        insurance=patient_fields.get("insurance"),  # type: ignore[arg-type]
        allergies_text=patient_fields.get("allergies_text"),  # type: ignore[arg-type]
        as_of_date=patient_fields.get("as_of_date"),  # type: ignore[arg-type]
        medications=medications,
        lab_trends=lab_trends,
        care_gaps=care_gaps,
        source_document_id=source_document_id,
    )


__all__ = [
    "WorkbookParseError",
    "parse_workbook",
]
